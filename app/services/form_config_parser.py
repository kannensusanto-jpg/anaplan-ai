import openpyxl
from io import BytesIO


def parse_form_config_template(data: bytes) -> dict:
    """
    Parse a filled Form Config Template and return a dict of fields
    suitable for updating a FormConfig record. Only keys with non-empty
    values are included, so callers can safely apply with setattr.
    """
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    if "Form Setup" not in wb.sheetnames:
        raise ValueError("Missing 'Form Setup' sheet in uploaded file")

    ws = wb["Form Setup"]
    raw: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[0] or "").strip()
        val = row[1]
        if not key or key.startswith("—"):
            continue
        raw[key] = val

    # DIM_ROLE_* → dimension_roles dict  {"account": "Accounts", ...}
    dimension_roles: dict[str, str] = {}
    for key, val in raw.items():
        if key.startswith("DIM_ROLE_") and val is not None and str(val).strip():
            role = key[len("DIM_ROLE_"):].lower()
            dimension_roles[role] = str(val).strip()

    # PAGE_SELECTOR_* → page_selectors dict  {"Product": "All Products", ...}
    page_selectors: dict[str, str] = {}
    for key, val in raw.items():
        if key.startswith("PAGE_SELECTOR_") and val is not None:
            val_str = str(val).strip()
            if "=" in val_str:
                dim, member = val_str.split("=", 1)
                page_selectors[dim.strip()] = member.strip()

    result: dict = {}

    def _str(k: str) -> str | None:
        v = raw.get(k)
        return str(v).strip() if v is not None and str(v).strip() else None

    def _int(k: str) -> int | None:
        v = raw.get(k)
        try:
            return int(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    for field, getter in [
        ("form_name",             lambda: _str("FORM_NAME")),
        ("profile_name",          lambda: _str("PROFILE_NAME")),
        ("actual_version_member", lambda: _str("ACTUAL_VERSION")),
        ("budget_version_member", lambda: _str("BUDGET_VERSION")),
        ("view_id",               lambda: _str("VIEW_ID")),
        ("import_action_id",      lambda: _str("IMPORT_ACTION_ID")),
        ("header_rows",           lambda: _int("HEADER_ROWS")),
        ("account_col",           lambda: _int("ACCOUNT_COL")),
    ]:
        v = getter()
        if v is not None:
            result[field] = v

    if dimension_roles:
        result["dimension_roles"] = dimension_roles
    if page_selectors:
        result["page_selectors"] = page_selectors

    return result
