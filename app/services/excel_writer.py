import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill

AI_HEADER    = "AI Commentary"
FILL_AI      = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_SKIPPED = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")


def write_commentary_to_excel(
    original_bytes: bytes,
    commentary: dict[str, str],
    skipped: list[tuple[dict, str]],
) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(original_bytes))
    ws = wb["Data"]

    headers = [cell.value for cell in ws[1]]
    ai_col  = headers.index(AI_HEADER) + 1 if AI_HEADER in headers else len(headers) + 1
    if AI_HEADER not in headers:
        cell      = ws.cell(row=1, column=ai_col, value=AI_HEADER)
        cell.font = Font(bold=True, color="1B5E20")

    col = {name: i for i, name in enumerate(headers)}
    row_index: dict[str, int] = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[col["Account"]]:
            mid = f"{row[col['Account']]}|{row[col['Cost Center']]}|{row[col['Time Period']]}"
            row_index[mid] = row_num

    for member_id, text in commentary.items():
        if member_id in row_index:
            cell      = ws.cell(row=row_index[member_id], column=ai_col, value=text)
            cell.fill = FILL_AI

    skipped_ids = {r["member_id"] for r, _ in skipped}
    for member_id in skipped_ids:
        if member_id in row_index:
            ws.cell(row=row_index[member_id], column=ai_col).fill = FILL_SKIPPED

    if skipped:
        if "Skipped" in wb.sheetnames:
            del wb["Skipped"]
        ws_skip = wb.create_sheet("Skipped")
        ws_skip.append(["Account", "Cost Center", "Time Period", "Variance $", "Reason"])
        ws_skip[1][0].font = Font(bold=True)
        for r, reason in skipped:
            ws_skip.append([r["account"], r["cost_center"], r["time_period"], r["variance_dollars"], reason])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_preview_excel(preview: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commentary"

    headers     = ["Account", "Cost Center", "Time Period", "Actual", "Budget", "Variance $", "Variance %", "AI Commentary"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell      = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    red_font   = Font(color="B91C1C")
    green_font = Font(color="15803D")
    bold_font  = Font(bold=True)

    for row in preview["rows"]:
        var = row["variance_dollars"]
        pct = row["variance_pct"]
        ws.append([
            row["account"], row["cost_center"], row["time_period"],
            row["actual"], row["budget"], var, round(pct * 100, 2), row["commentary"],
        ])
        r          = ws.max_row
        color_font = red_font if var > 0 else green_font
        for c in range(6, 8):
            ws.cell(row=r, column=c).font = color_font
        if row.get("is_rollup"):
            ws.cell(row=r, column=1).font = bold_font

    ws.column_dimensions["H"].width = 60
    ws.column_dimensions["A"].width = 24

    if preview["skipped"]:
        ws_skip = wb.create_sheet("Skipped")
        ws_skip.append(["Account", "Cost Center", "Time Period", "Variance $", "Reason"])
        ws_skip[1][0].font = Font(bold=True)
        for s in preview["skipped"]:
            ws_skip.append([s["account"], s["cost_center"], s["time_period"], s["variance_dollars"], s["reason"]])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
