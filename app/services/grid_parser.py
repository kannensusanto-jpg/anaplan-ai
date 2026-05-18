"""
Parses Anaplan-style grid Excel exports into the flat row format the
generation pipeline expects.

Supported grid shapes
─────────────────────
1. Single-header (header_rows=1):
       Account | Actual | Budget | Prior Year | Account Type | ...

2. Two-row pivot header (header_rows=2):
       Account | FY26 Q1 | FY26 Q1 | FY26 Q2 | FY26 Q2
               | Actual  | Budget  | Actual  | Budget

3. Multi-tab: each sheet represents a different entity (department /
   cost-centre). page_selectors or the sheet name supplies the entity label.

Hierarchy detection
───────────────────
Leading-space indentation on the account name column identifies
parent-child relationships.  Total/rollup rows have LESS indentation than
their children and appear AFTER them in the grid, so hierarchy is resolved
in a second pass (look-ahead for the nearest row with lower indent).

column_mapping format (stored in FormConfig.column_mapping, keys are
0-based column indices as strings):
    {
      "2": {"role": "actual",       "time": "FY26 Q1"},
      "3": {"role": "budget",       "time": "FY26 Q1"},
      "4": {"role": "prior_actual", "time": "FY25 Q4"},
      "5": {"role": "account_type"},
      "6": {"role": "commentary"},
      "7": {"role": "parent_member"}
    }
When column_mapping is empty the parser auto-detects roles from headers.
"""

import re
from dataclasses import dataclass, field
from io import BytesIO

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.form_config import FormConfig


# ── Column role constants ────────────────────────────────────────────────────

ACTUAL       = "actual"
BUDGET       = "budget"
PRIOR_ACTUAL = "prior_actual"
ACCOUNT_TYPE = "account_type"
COMMENTARY   = "commentary"
PARENT       = "parent_member"

# Header-text → role mapping for auto-detection
_AUTO_ROLES: dict[str, str] = {
    "actual":               ACTUAL,
    "actuals":              ACTUAL,
    "act":                  ACTUAL,
    "budget":               BUDGET,
    "plan":                 BUDGET,
    "forecast":             BUDGET,
    "fcast":                BUDGET,
    "prior actual":         PRIOR_ACTUAL,
    "prior period actual":  PRIOR_ACTUAL,
    "prior period":         PRIOR_ACTUAL,
    "prior year":           PRIOR_ACTUAL,
    "py":                   PRIOR_ACTUAL,
    "ly":                   PRIOR_ACTUAL,
    "last year":            PRIOR_ACTUAL,
    "account type":         ACCOUNT_TYPE,
    "type":                 ACCOUNT_TYPE,
    "commentary":           COMMENTARY,
    "human commentary":     COMMENTARY,
    "notes":                COMMENTARY,
    "analyst notes":        COMMENTARY,
    "parent":               PARENT,
    "parent member":        PARENT,
}


@dataclass
class _ColSpec:
    role: str
    time: str = ""   # time-period label for this column, e.g. "FY26 Q1"


# ── Public API ───────────────────────────────────────────────────────────────

def parse_grid(
    file_bytes: bytes,
    form_config: FormConfig,
) -> tuple[list[dict], dict[str, list[str]]]:
    """
    Parse an Anaplan-style grid workbook.

    Returns:
        rows      — list of dicts compatible with generate_commentary()
        hierarchy — {parent_member_id: [child_member_id, ...]}
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    all_rows: list[dict] = []
    all_hier: dict[str, list[str]] = {}

    for sheet_name in wb.sheetnames:
        if _should_skip_sheet(sheet_name, form_config):
            continue
        ws      = wb[sheet_name]
        entity  = _resolve_entity(sheet_name, form_config)
        default_time = _resolve_default_time(form_config)
        specs   = _resolve_col_specs(ws, form_config, default_time)

        if not specs:
            continue  # no recognisable data columns, skip silently

        rows, hier = _parse_sheet(ws, entity, specs, form_config)

        all_rows.extend(rows)
        for parent_id, children in hier.items():
            all_hier.setdefault(parent_id, []).extend(children)

    return all_rows, all_hier


# ── Sheet-level helpers ──────────────────────────────────────────────────────

def _should_skip_sheet(sheet_name: str, fc: FormConfig) -> bool:
    """Skip sheets that are clearly metadata (Config, Instructions, etc.)."""
    skip_names = {"config", "instructions", "skipped", "form setup",
                  "config overrides", "form config"}
    return sheet_name.lower() in skip_names


def _resolve_entity(sheet_name: str, fc: FormConfig) -> str:
    """Determine entity (cost-centre) label for this sheet."""
    if fc.page_selectors:
        entity_dim = (fc.dimension_roles or {}).get("entity", "")
        if entity_dim and entity_dim in fc.page_selectors:
            return fc.page_selectors[entity_dim]
        return next(iter(fc.page_selectors.values()))
    return sheet_name


def _resolve_default_time(fc: FormConfig) -> str:
    """Return a time-period label if it's fixed via page selectors."""
    if fc.page_selectors:
        time_dim = (fc.dimension_roles or {}).get("time", "")
        if time_dim and time_dim in fc.page_selectors:
            return fc.page_selectors[time_dim]
    return ""


def _resolve_col_specs(
    ws: Worksheet, fc: FormConfig, default_time: str
) -> dict[int, _ColSpec]:
    """Build 0-based column index → _ColSpec map."""
    if fc.column_mapping:
        return {
            int(k): _ColSpec(
                role=v.get("role", ""),
                time=v.get("time", default_time),
            )
            for k, v in fc.column_mapping.items()
            if v.get("role")
        }
    return _auto_detect_specs(ws, fc.header_rows, fc.account_col, default_time)


def _auto_detect_specs(
    ws: Worksheet, header_rows: int, account_col: int, default_time: str
) -> dict[int, _ColSpec]:
    """Infer column roles from the header row(s)."""
    specs: dict[int, _ColSpec] = {}

    def _cell(r: int, c: int):  # 1-based
        return ws.cell(row=r, column=c + 1).value

    if header_rows == 1:
        for col_idx in range(ws.max_column):
            if col_idx == account_col:
                continue
            h = _cell(1, col_idx)
            if h is None:
                continue
            role = _AUTO_ROLES.get(str(h).strip().lower())
            if role:
                specs[col_idx] = _ColSpec(role=role, time=default_time)

    elif header_rows >= 2:
        # Row 1 carries the time-period context; row 2 the version label.
        # Merge-span cells leave None in subsequent columns — propagate forward.
        time_ctx = default_time
        for col_idx in range(ws.max_column):
            if col_idx == account_col:
                continue
            r1 = _cell(1, col_idx)
            r2 = _cell(2, col_idx)
            if r1 is not None:
                time_ctx = str(r1).strip()
            version = str(r2).strip().lower() if r2 else ""
            role = _AUTO_ROLES.get(version)
            if role:
                specs[col_idx] = _ColSpec(role=role, time=time_ctx)

    return specs


# ── Row-level parsing ────────────────────────────────────────────────────────

def _parse_sheet(
    ws: Worksheet,
    entity: str,
    specs: dict[int, _ColSpec],
    fc: FormConfig,
) -> tuple[list[dict], dict[str, list[str]]]:
    data_start  = fc.header_rows + 1  # 1-based
    acc_col     = fc.account_col       # 0-based
    ent_col     = fc.entity_col        # 0-based or None

    actual_ver  = (fc.actual_version_member or "Actual").lower()
    budget_ver  = (fc.budget_version_member or "Budget").lower()

    # Pass 1: collect raw entries with indent level
    raw: list[dict] = []   # {indent, account, entity, time_period→{actual, budget, …}}

    for sheet_row_idx, row_cells in enumerate(
        ws.iter_rows(min_row=data_start), start=data_start
    ):
        acc_cell = row_cells[acc_col] if acc_col < len(row_cells) else None
        if acc_cell is None or acc_cell.value is None:
            continue

        raw_val   = str(acc_cell.value)
        stripped  = raw_val.lstrip()
        # Prefer Excel indent attribute (survives when leading-space is used)
        excel_indent = getattr(acc_cell.alignment, "indent", 0) or 0
        space_indent = len(raw_val) - len(stripped)
        indent_level = excel_indent if excel_indent else space_indent

        account_name = stripped.strip()
        if not account_name:
            continue

        time_data: dict[str, dict] = {}
        for col_idx, spec in specs.items():
            if col_idx >= len(row_cells):
                continue
            val = row_cells[col_idx].value
            td  = time_data.setdefault(spec.time, {})
            _accumulate(td, spec.role, val)

        if not time_data:
            continue

        raw.append({
            "indent":      indent_level,
            "account":     account_name,
            "time_data":   time_data,
            "row_idx":     sheet_row_idx,
        })

    # Pass 2: build parent map from indentation (look-ahead for lower indent)
    parent_map = _build_parent_map(raw)

    # Pass 3: emit final row dicts
    rows: list[dict] = []
    hierarchy: dict[str, list[str]] = {}

    for entry in raw:
        account = entry["account"]
        for time_period, td in entry["time_data"].items():
            actual  = td.get(ACTUAL, 0.0)
            budget  = td.get(BUDGET, 0.0)
            member_id = _make_id(account, entity, time_period)

            parent_account = parent_map.get(entry["account"])
            parent_id = _make_id(parent_account, entity, time_period) if parent_account else None
            if parent_id:
                hierarchy.setdefault(parent_id, []).append(member_id)

            rows.append({
                "member_id":        member_id,
                "account":          account,
                "cost_center":      entity,
                "time_period":      time_period or "N/A",
                "actual":           actual,
                "budget":           budget,
                "variance_dollars": actual - budget,
                "variance_pct":     (actual - budget) / budget if budget else 0.0,
                "prior_actual":     td.get(PRIOR_ACTUAL, 0.0),
                "account_type":     td.get(ACCOUNT_TYPE, "expense"),
                "human_commentary": td.get(COMMENTARY) or None,
                "parent_member_id": parent_id,
            })

    return rows, hierarchy


def _accumulate(td: dict, role: str, val) -> None:
    if role == ACTUAL:
        td[ACTUAL]       = _to_float(val)
    elif role == BUDGET:
        td[BUDGET]       = _to_float(val)
    elif role == PRIOR_ACTUAL:
        td[PRIOR_ACTUAL] = _to_float(val)
    elif role == ACCOUNT_TYPE:
        td[ACCOUNT_TYPE] = _to_account_type(val)
    elif role == COMMENTARY:
        td[COMMENTARY]   = str(val).strip() if val else None
    elif role == PARENT:
        if val:
            td[PARENT] = str(val).strip()


def _build_parent_map(raw: list[dict]) -> dict[str, str]:
    """
    For each entry, its parent is the NEXT entry with strictly lower indent.
    (In Anaplan grids, totals appear after their children.)
    """
    n = len(raw)
    parent_map: dict[str, str] = {}
    for i, entry in enumerate(raw):
        for j in range(i + 1, n):
            if raw[j]["indent"] < entry["indent"]:
                parent_map[entry["account"]] = raw[j]["account"]
                break
    return parent_map


def _make_id(account: str | None, entity: str, time_period: str) -> str | None:
    if not account:
        return None
    return f"{account}|{entity}|{time_period or 'N/A'}"


def _to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = re.sub(r"[^\d.\-]", "", str(val))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _to_account_type(val) -> str:
    if not val:
        return "expense"
    v = str(val).strip().lower()
    return "revenue" if ("revenue" in v or "income" in v) else "expense"
