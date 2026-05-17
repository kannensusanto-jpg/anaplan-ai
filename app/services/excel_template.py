import openpyxl
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill


def generate_template() -> bytes:
    wb = openpyxl.Workbook()
    _build_data_sheet(wb)
    _build_config_sheet(wb)
    _build_instructions_sheet(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_data_sheet(wb: openpyxl.Workbook) -> None:
    ws      = wb.active
    ws.title = "Data"
    headers = [
        "Account", "Cost Center", "Time Period",
        "Actual", "Budget", "Prior Period Actual",
        "Account Type", "Parent Member", "Human Commentary",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(wrap_text=True)
    samples = [
        # Account               Cost Ctr   Period     Actual      Budget    Prior Act  Acct Type   Parent          Human Commentary
        ["Salaries",            "Dept A",  "FY26 Q1", 1_150_000, 1_000_000,  975_000, "expense",  "Total Payroll", "New hires onboarded in Jan"],
        ["Benefits",            "Dept A",  "FY26 Q1",   380_000,   300_000,  290_000, "expense",  "Total Payroll", ""],
        ["Total Payroll",       "Dept A",  "FY26 Q1", 1_530_000, 1_300_000, 1_265_000, "expense", "",              ""],
        ["Travel & Expense",    "Dept A",  "FY26 Q1",   112_000,    60_000,   55_000, "expense",  "Total OpEx",    "Q1 sales kickoff event"],
        ["Software Licenses",   "Dept A",  "FY26 Q1",   210_000,   175_000,  170_000, "expense",  "Total OpEx",    ""],
        ["Consulting Fees",     "Dept A",  "FY26 Q1",    42_000,    50_000,   48_000, "expense",  "Total OpEx",    ""],
        ["Total OpEx",          "Dept A",  "FY26 Q1",   364_000,   285_000,  273_000, "expense",  "",              ""],
        ["Product Revenue",     "Dept A",  "FY26 Q1", 4_250_000, 4_000_000, 3_800_000, "revenue", "Total Revenue", ""],
        ["Services Revenue",    "Dept A",  "FY26 Q1",   680_000,   750_000,  710_000, "revenue",  "Total Revenue", "Two enterprise deals pushed to Q2"],
        ["Total Revenue",       "Dept A",  "FY26 Q1", 4_930_000, 4_750_000, 4_510_000, "revenue", "",             ""],
    ]
    for row in samples:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["I"].width = 40


def _build_config_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Config")
    ws.append(["Key", "Value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    config_rows = [
        ("TONE",                    "concise and direct"),
        ("MATERIALITY_THRESHOLD_$", 50_000),
        ("MATERIALITY_THRESHOLD_%", 0.05),
        ("FOCUS_AREAS",             "headcount, T&E, software licenses"),
        ("ROLLUP_LEVELS",           "Department, Division"),
        ("SUPPRESS_FAVORABLE",      "FALSE"),
        ("PRIOR_PERIOD_CONTEXT",    "TRUE"),
    ]
    for row in config_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40


def _build_instructions_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    lines = [
        ("AI Commentary — Upload Template",                                            True),
        ("",                                                                           False),
        ("DATA SHEET COLUMNS",                                                         True),
        ("Account — account name as it appears in your EPM system",                   False),
        ("Cost Center — cost center, department, or entity",                          False),
        ("Time Period — e.g. FY26 Q1, Jan-2026, FY2026",                              False),
        ("Actual — actual result (numeric)",                                           False),
        ("Budget — budget or plan value (numeric)",                                    False),
        ("Prior Period Actual — optional, used for trend context",                    False),
        ("Account Type — expense or revenue (determines favorable direction)",         False),
        ("Parent Member — optional, rollup parent for synthesis commentary",           False),
        ("Human Commentary — optional, analyst notes visible to AI as context",       False),
        ("",                                                                           False),
        ("CONFIG SHEET",                                                               True),
        ("Edit the Value column to configure AI behavior for this upload.",            False),
        ("",                                                                           False),
        ("OUTPUT",                                                                     True),
        ("An AI Commentary column is added to your Data sheet (green = generated).",  False),
        ("Rows below materiality threshold are highlighted yellow.",                   False),
        ("A Skipped sheet lists filtered rows and the reason each was skipped.",       False),
    ]
    for text, bold in lines:
        ws.append([text])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 70
