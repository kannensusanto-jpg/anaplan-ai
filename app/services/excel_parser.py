import openpyxl
from io import BytesIO

from app.services.config import ClientConfig

REQUIRED_COLUMNS = {"Account", "Cost Center", "Time Period", "Actual", "Budget"}


def parse_excel(file_bytes: bytes) -> tuple[ClientConfig, list[dict], dict[str, list[str]]]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    config = _parse_config(wb)
    rows, hierarchy = _parse_data(wb)
    return config, rows, hierarchy


def _parse_config(wb: openpyxl.Workbook) -> ClientConfig:
    if "Config" not in wb.sheetnames:
        return _default_config()
    ws  = wb["Config"]
    cfg = {
        str(row[0]).strip(): str(row[1]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0] and row[1] is not None
    }
    return ClientConfig(
        tone=cfg.get("TONE", "concise and direct"),
        materiality_dollars=float(cfg.get("MATERIALITY_THRESHOLD_$", 50_000)),
        materiality_pct=float(cfg.get("MATERIALITY_THRESHOLD_%", 0.05)),
        focus_areas=[s.strip() for s in cfg.get("FOCUS_AREAS", "").split(",") if s.strip()],
        rollup_levels=[s.strip() for s in cfg.get("ROLLUP_LEVELS", "").split(",") if s.strip()],
        suppress_favorable=cfg.get("SUPPRESS_FAVORABLE", "FALSE").upper() == "TRUE",
        prior_period_context=cfg.get("PRIOR_PERIOD_CONTEXT", "TRUE").upper() == "TRUE",
    )


def _parse_data(wb: openpyxl.Workbook) -> tuple[list[dict], dict[str, list[str]]]:
    if "Data" not in wb.sheetnames:
        raise ValueError("File must contain a 'Data' sheet")

    ws      = wb["Data"]
    headers = [cell.value for cell in ws[1]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(f"Data sheet missing required columns: {missing}")

    col  = {name: i for i, name in enumerate(headers)}
    rows, hierarchy = [], {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[col["Account"]]:
            continue

        actual      = float(row[col["Actual"]] or 0)
        budget      = float(row[col["Budget"]] or 0)
        account     = str(row[col["Account"]])
        cost_center = str(row[col["Cost Center"]])
        time_period = str(row[col["Time Period"]])
        member_id   = f"{account}|{cost_center}|{time_period}"

        parent_raw = row[col["Parent Member"]] if "Parent Member" in col else None
        parent_id  = None
        if parent_raw:
            parent_id = f"{parent_raw}|{cost_center}|{time_period}"
            hierarchy.setdefault(parent_id, []).append(member_id)

        rows.append({
            "member_id":        member_id,
            "row_num":          row_num,
            "account":          account,
            "cost_center":      cost_center,
            "time_period":      time_period,
            "actual":           actual,
            "budget":           budget,
            "variance_dollars": actual - budget,
            "variance_pct":     (actual - budget) / budget if budget else 0,
            "prior_actual":     float(row[col["Prior Period Actual"]] or 0) if "Prior Period Actual" in col else 0,
            "account_type":     str(row[col["Account Type"]] or "expense").lower() if "Account Type" in col else "expense",
            "human_commentary": str(row[col["Human Commentary"]]) if "Human Commentary" in col and row[col["Human Commentary"]] else None,
            "parent_member_id": parent_id,
        })

    return rows, hierarchy


def _default_config() -> ClientConfig:
    return ClientConfig(
        tone="concise and direct",
        materiality_dollars=50_000,
        materiality_pct=0.05,
        focus_areas=[],
        rollup_levels=[],
        suppress_favorable=False,
        prior_period_context=True,
    )
