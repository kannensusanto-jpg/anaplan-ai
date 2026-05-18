import openpyxl
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill


def generate_form_config_template() -> bytes:
    wb = openpyxl.Workbook()
    _build_form_setup_sheet(wb)
    _build_overrides_sheet(wb)
    _build_instructions_sheet(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hdr(ws, text: str) -> None:
    ws.append([text])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)


def _row(ws, key: str, value, note: str = "") -> None:
    ws.append([key, value, note])


def _build_form_setup_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Form Setup"

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(["Setting", "Value", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font

    rows = [
        # ── Identity ───────────────────────────────────────────────────────
        ("FORM_ID",              "my-pl-dept-a",      "Unique slug — no spaces (used as identifier)"),
        ("FORM_NAME",            "P&L — Dept A",      "Human-readable display name"),
        ("FORM_SOURCE",          "excel",              '"excel" or "anaplan"'),
        ("PROFILE_NAME",         "P&L",                "Built-in: P&L, Headcount, CapEx, Balance Sheet, Cash Flow — or a custom name"),

        # ── Dimension roles ─────────────────────────────────────────────────
        ("— DIMENSION ROLES —",  "",                   "Map each semantic role to the Anaplan dimension name on this form"),
        ("DIM_ROLE_ACCOUNT",     "Account",            "Dimension holding account/line-item members (the row dimension)"),
        ("DIM_ROLE_TIME",        "Periods",            "Time dimension name"),
        ("DIM_ROLE_VERSION",     "Versions",           "Version dimension (Actual vs Budget)"),
        ("DIM_ROLE_ENTITY",      "Cost Centre",        "Entity / cost-centre dimension"),
        ("DIM_ROLE_COMMENTARY",  "Commentary",         "Dimension containing the AI Commentary member (write-back target)"),
        ("DIM_ROLE_PRODUCT",     "",                   "Optional: any additional dimension — add more DIM_ROLE_* rows as needed"),
        ("ACTUAL_VERSION",       "Actual",             "Member name for Actuals in the version dimension"),
        ("BUDGET_VERSION",       "Budget",             "Member name for Budget / Plan"),

        # ── Page selectors ──────────────────────────────────────────────────
        ("— PAGE SELECTORS —",   "",                   "Fixed dimension values for context dimensions not represented in rows"),
        ("PAGE_SELECTOR_1",      "Product=All Products", '"DimensionName=MemberName" — add more PAGE_SELECTOR_N rows as needed'),
        ("PAGE_SELECTOR_2",      "",                   "Additional page selector"),

        # ── Excel grid specifics ────────────────────────────────────────────
        ("— EXCEL GRID —",       "",                   ""),
        ("HEADER_ROWS",          1,                    "Number of header rows in the Excel grid (1 or 2)"),
        ("ACCOUNT_COL",          0,                    "0-based column index of the account / row label column"),

        # ── Anaplan specifics ───────────────────────────────────────────────
        ("— ANAPLAN —",          "",                   ""),
        ("VIEW_ID",              "",                   "Anaplan view ID (leave blank for Excel source)"),
        ("IMPORT_ACTION_ID",     "",                   "Anaplan import action ID for writing commentary back"),
    ]

    for key, value, note in rows:
        ws.append([key, value, note])
        if key.startswith("—"):
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55

    # Highlight example values in light blue so finance knows to edit them
    edit_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.fill = edit_fill


def _build_overrides_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Config Overrides")

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(["Override Key", "Value", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font

    ws.append(["TONE",                    "",     "Leave blank to use profile default. e.g. 'executive summary'"])
    ws.append(["MATERIALITY_THRESHOLD_$", "",     "Minimum absolute variance to generate commentary"])
    ws.append(["MATERIALITY_THRESHOLD_%", "",     "Minimum % variance (e.g. 0.05 = 5%)"])
    ws.append(["FOCUS_AREAS",             "",     "Comma-separated focus areas to emphasize"])
    ws.append(["ROLLUP_LEVELS",           "",     "Comma-separated rollup levels for synthesis"])
    ws.append(["SUPPRESS_FAVORABLE",      "",     '"TRUE" to skip favorable variances'])
    ws.append(["PRIOR_PERIOD_CONTEXT",    "",     '"FALSE" to omit prior period trend language'])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55


def _build_instructions_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Instructions")

    lines = [
        ("Form Configuration Template",                                                          True),
        ("",                                                                                     False),
        ("HOW TO USE",                                                                           True),
        ("1. Fill in the 'Form Setup' sheet. Only edit the Value column (blue cells).",          False),
        ("2. Optionally override profile defaults in 'Config Overrides' (leave blank to skip).", False),
        ("3. Upload this file via the Form Config upload endpoint.",                             False),
        ("4. The form will appear in the Form Picker on the main dashboard.",                    False),
        ("",                                                                                     False),
        ("FORM SOURCE",                                                                          True),
        ("excel  — upload an Excel grid that mirrors your Anaplan view layout.",                 False),
        ("anaplan — the tool reads the view directly using the Anaplan API.",                    False),
        ("",                                                                                     False),
        ("PROFILE NAMES (built-in)",                                                             True),
        ("P&L           — income statement with revenue and expense commentary",                 False),
        ("Headcount     — headcount and payroll focus, lower materiality threshold",             False),
        ("CapEx         — capital expenditure, projects, depreciation",                          False),
        ("Balance Sheet — assets, liabilities, equity",                                          False),
        ("Cash Flow     — operating, investing, financing activities",                           False),
        ("",                                                                                     False),
        ("DIMENSION ROLES",                                                                      True),
        ("DIM_ROLE_* rows map semantic roles to Anaplan dimension names.",                       False),
        ("Known roles: account, time, version, entity, commentary.",                             False),
        ("Add more DIM_ROLE_* rows for any extra dimensions on the form (e.g. DIM_ROLE_PRODUCT).", False),
        ("",                                                                                     False),
        ("PAGE SELECTORS",                                                                       True),
        ("Use PAGE_SELECTOR_N rows for dimensions with a fixed value not in the row axis.",      False),
        ("Format: 'DimensionName=MemberName' — e.g. 'Product=All Products'.",                   False),
    ]

    for text, bold in lines:
        ws.append([text])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 75
